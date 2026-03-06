from openpyxl import load_workbook
from pathlib import Path

# --- НАСТРОЙКИ ---
FILE_NAME = 'map1.xlsx'
FOLDER_NAME = 'maps'
# -----------------

# 1. Определяем путь к файлу
script_dir = Path(__file__).parent
excel_path = script_dir.parent / FOLDER_NAME / FILE_NAME

if not excel_path.exists():
    print(f"❌ Ошибка: Файл не найден по пути: {excel_path}")
    exit()

print(f"✅ Загрузка файла: {excel_path}")
wb = load_workbook(excel_path, data_only=True)
ws = wb.active


# --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ---
def get_cell_color_hex(cell):
    """Возвращает HEX цвет ячейки или None, если цвета нет"""
    fill = cell.fill
    if fill.fill_type is None:
        return None
    fg_color = fill.fgColor
    if fg_color.type != 'rgb':
        return None
    # Убираем первые 2 символа (прозрачность)
    return fg_color.rgb[2:]


# --- 2. НАХОДИМ ГРАНИЦЫ ЧЕРНОГО ПРЯМОУГОЛЬНИКА ---
# Черный цвет в HEX обычно '000000'
BLACK_COLOR = '000000'
START_ROW = 1
START_COL = 1

# Проверяем, что стартовая ячейка действительно черная
first_cell_color = get_cell_color_hex(ws.cell(row=START_ROW, column=START_COL))
if first_cell_color != BLACK_COLOR:
    print(f"❌ Ошибка: Ячейка A1 не черная (цвет: {first_cell_color})")
    print("   Скрипт ожидает, что карта начинается с черного прямоугольника.")
    exit()

print("🔍 Поиск границ черного прямоугольника...")

# Находим правую границу (ширину)
# Идем вправо по первой строке, пока цвет черный
width = 0
for col in range(START_COL, ws.max_column + 1):
    color = get_cell_color_hex(ws.cell(row=START_ROW, column=col))
    if color == BLACK_COLOR:
        width += 1
    else:
        break

# Находим нижнюю границу (высоту)
# Идем вниз по первому столбцу, пока цвет черный
height = 0
for row in range(START_ROW, ws.max_row + 1):
    color = get_cell_color_hex(ws.cell(row=row, column=START_COL))
    if color == BLACK_COLOR:
        height += 1
    else:
        break

print(f"✅ Найдено: Ширина = {width}, Высота = {height}")

# --- 3. ИЗВЛЕКАЕМ ОСНОВНУЮ КАРТУ (ВНУТРЕННОСТЬ ПРЯМОУГОЛЬНИКА) ---
print(f"⏳ Чтение данных карты ({width}x{height})...")

main_map = []

for row in range(START_ROW, START_ROW + height):
    row_data = []
    for col in range(START_COL, START_COL + width):
        cell = ws.cell(row=row, column=col)
        color = get_cell_color_hex(cell)

        # Сохраняем цвет (если None - значит без заливки)
        row_data.append(color)
    main_map.append(row_data)


for row in main_map:
    print(row)
# --- 4. ВЫВОД И СОХРАНЕНИЕ ---

# Пример: вывести размер полученной карты
print(f"📊 Размер основной карты: {len(main_map)} строк * {len(main_map[0])} столбцов")
